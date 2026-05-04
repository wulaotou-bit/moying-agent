"""
墨影数据监控分析模块
功能：
1. 解析番茄作家后台数据
2. 生成带图表的Excel报告
3. 基于数据给出写作优化建议
4. 历史数据趋势分析
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from config import Config
from utils.logger import log
from datetime import datetime
import os

class DataAnalyzer:
    def __init__(self):
        self.report_dir = Config.OUTPUT_DIR / "数据报告"
        self.report_dir.mkdir(exist_ok=True)
        self.history_path = self.report_dir / "历史数据.xlsx"
        log.success("✅ 数据监控分析模块初始化完成")

    def parse_raw_data(self, raw_data: dict) -> pd.DataFrame:
        """解析爬虫爬取的原始数据"""
        try:
            df = pd.DataFrame([raw_data])
            df["时间"] = pd.to_datetime(df["时间"])
            df.set_index("时间", inplace=True)
            # 转换数值类型（容错处理）
            for col in ["阅读量", "收藏量"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            for col in ["追读率", "完读率"]:
                if col in df.columns:
                    df[col] = df[col].str.replace("%", "", regex=False)
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            log.info("原始数据解析成功")
            return df
        except Exception as e:
            log.error(f"原始数据解析失败: {e}")
            return pd.DataFrame()

    def append_to_history(self, df: pd.DataFrame) -> pd.DataFrame:
        """追加数据到历史记录，自动去重"""
        if self.history_path.exists():
            try:
                history_df = pd.read_excel(self.history_path, index_col=0, parse_dates=True)
                full_df = pd.concat([history_df, df])
                full_df = full_df[~full_df.index.duplicated(keep='last')]
                full_df = full_df.sort_index()
                log.info(f"历史数据追加成功，共{len(full_df)}条记录")
            except Exception as e:
                log.error(f"历史数据加载失败: {e}，使用新数据")
                full_df = df
        else:
            full_df = df
            log.info("首次生成历史数据文件")
        
        # 保存历史数据
        full_df.to_excel(self.history_path)
        return full_df

    def generate_excel_report(self, full_df: pd.DataFrame) -> str:
        """生成带图表的美化Excel报告"""
        try:
            report_name = f"数据报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_path = self.report_dir / report_name

            # 保存基础数据
            full_df.to_excel(report_path, sheet_name="数据概览")

            # 加载工作簿进行美化
            wb = load_workbook(report_path)
            ws = wb["数据概览"]

            # 美化表头
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column] = adjusted_width

            # 根据列名动态获取列号（防止列顺序变化）
            col_map = {cell.value: i+1 for i, cell in enumerate(ws[1])}
            read_col = col_map.get("阅读量", 2)
            collect_col = col_map.get("收藏量", 3)
            zhuidu_col = col_map.get("追读率", 4)
            wandu_col = col_map.get("完读率", 5)

            # 添加阅读量/收藏量折线图
            chart1 = LineChart()
            chart1.title = "阅读量/收藏量趋势"
            chart1.y_axis.title = "数值"
            chart1.x_axis.title = "时间"
            chart1.style = 10
            chart1.height = 10
            chart1.width = 15
            data1 = Reference(ws, min_col=read_col, min_row=1, max_col=collect_col, max_row=ws.max_row)
            cats1 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)
            ws.add_chart(chart1, "F2")

            # 添加追读率/完读率折线图
            chart2 = LineChart()
            chart2.title = "追读率/完读率趋势"
            chart2.y_axis.title = "百分比(%)"
            chart2.x_axis.title = "时间"
            chart2.style = 11
            chart2.height = 10
            chart2.width = 15
            data2 = Reference(ws, min_col=zhuidu_col, min_row=1, max_col=wandu_col, max_row=ws.max_row)
            cats2 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            ws.add_chart(chart2, "F20")

            # 保存报告
            wb.save(report_path)
            log.success(f"Excel报告生成成功: {report_path}")
            return str(report_path)
        except Exception as e:
            log.error(f"Excel报告生成失败: {e}")
            return ""

    def generate_writing_suggestions(self, full_df: pd.DataFrame) -> str:
        """基于数据生成专业写作优化建议"""
        try:
            if full_df.empty or len(full_df) < 1:
                return "暂无足够数据，无法生成建议"
            
            latest = full_df.iloc[-1]
            suggestions = []

            # 追读率分析（番茄最重要指标）
            if "追读率" in latest:
                zhuidu = latest["追读率"]
                if zhuidu < 15:
                    suggestions.append("⚠️ 【严重】追读率过低！请立即优化前3章开头，前300字必须进入冲突")
                elif zhuidu < 25:
                    suggestions.append("📝 追读率偏低，建议每章开头加悬念，中间增加爽点密度")
                elif zhuidu < 40:
                    suggestions.append("✅ 追读率良好，保持当前节奏")
                else:
                    suggestions.append("🌟 追读率优秀！可以适当增加更新频率")

            # 完读率分析
            if "完读率" in latest:
                wandu = latest["完读率"]
                if wandu < 8:
                    suggestions.append("⚠️ 完读率过低，建议控制单章字数在2000字左右，结尾必须留钩子")
                elif wandu < 15:
                    suggestions.append("📝 完读率一般，建议每章结尾加"下章预告"引导追读")
                else:
                    suggestions.append("✅ 完读率优秀，保持当前结尾风格")

            # 收藏增长分析
            if "收藏量" in latest and len(full_df) > 1:
                prev = full_df.iloc[-2]
                delta = latest["收藏量"] - prev["收藏量"]
                if delta < 5:
                    suggestions.append("⚠️ 收藏增长缓慢，建议优化作品简介和封面")
                elif delta < 20:
                    suggestions.append("📝 收藏增长一般，建议在章节末尾引导读者收藏")
                else:
                    suggestions.append("✅ 收藏增长优秀，继续保持")

            # 阅读量分析
            if "阅读量" in latest and len(full_df) > 1:
                prev_read = full_df.iloc[-2]["阅读量"]
                read_growth = (latest["阅读量"] - prev_read) / prev_read * 100 if prev_read > 0 else 0
                if read_growth < 5:
                    suggestions.append("📝 阅读量增长缓慢，建议多参与平台活动")
                elif read_growth > 20:
                    suggestions.append("🌟 阅读量爆发式增长！抓住机会加更")

            return "\n".join(suggestions) if suggestions else "✅ 数据表现全面优秀，无需调整"
        except Exception as e:
            log.error(f"写作建议生成失败: {e}")
            return "生成建议失败，请检查数据"

    def run_full_analysis(self, raw_data: dict) -> dict:
        """运行完整数据分析流程"""
        log.info("🚀 启动完整数据分析流程")
        df = self.parse_raw_data(raw_data)
        if df.empty:
            return {"report_path": "", "suggestions": "数据解析失败"}
        
        full_df = self.append_to_history(df)
        report_path = self.generate_excel_report(full_df)
        suggestions = self.generate_writing_suggestions(full_df)
        
        log.info("✅ 数据分析完成")
        return {
            "report_path": report_path,
            "suggestions": suggestions,
            "latest_data": raw_data
        }